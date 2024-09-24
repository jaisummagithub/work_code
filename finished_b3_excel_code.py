import os
import openpyxl as op
from openpyxl import Workbook
from datetime import datetime
from time import sleep as wait
from openpyxl.utils import get_column_letter
# Get the current date as a string
current_date = datetime.now().strftime('%d-%m-%Y %H.%M')

# Replace these file paths with your actual file paths 
jira_file_path = r'C:\Users\Jaa54335\Desktop\excel_code\Jira.xlsx'                  ## change this for your requirments
master_file_path = r'C:\Users\Jaa54335\Desktop\excel_code\b3_sheet.xlsx'            ## change this for your requirments
output_file_path = r'C:\Users\Jaa54335\Desktop\excel_code\summa_output_file.xlsx'   ## change this for your requirments

# Creates a new file if it doesnt exist
if not os.path.exists(output_file_path):# Check if the file exists
    # If the file doesn't exist, create an empty workbook using openpyxl
    workbook = Workbook()
    # Save the workbook
    workbook.save(output_file_path)
    print(f"Empty file created at: {output_file_path}")

# Load Jira file, master sheet, and create a new output file
jira_workbook = op.load_workbook(jira_file_path)
master_workbook = op.load_workbook(master_file_path)
output_workbook = op.load_workbook(output_file_path)

# Create a new sheet with the current date as its name
new_sheet = output_workbook.create_sheet(title=current_date)
output_workbook.move_sheet(new_sheet, offset=-output_workbook.index(new_sheet))  # Move the sheet to the beginning
output_workbook.active = output_workbook.sheetnames.index(current_date)  # Set the newly created sheet as active

# Get the first sheet in each workbook
jira_sheet = jira_workbook.active
master_sheet = master_workbook.active
output_sheet = output_workbook.active

# Initialize header index dictionaries
Jira_header_index = {'Issue key': 0, "Updated": 0, "Custom field (Responsible Team)": 0, "Summary": 0, "Status": 0,
                     "Issue id": 0, "Issue Type": 0, "Custom field (Preventive Action Category)": 0, "Project key": 0,
                     "Project name": 0, "Project type": 0, "Project lead": 0, "Project description": 0,
                     "Project url": 0, "Priority": 0, "Resolution": 0, "Assignee": 0, "Reporter": 0, "Creator": 0}

master_header_index = {'Issue key': 0, "Updated": 1, "Custom field (Responsible Team)": 0, "Summary": 0, "Status": 0,
                       "Issue id": 0, "Issue Type": 0, "Custom field (Preventive Action Category)": 0, "Project key": 0,
                       "Project name": 0, "Project type": 0, "Project lead": 0, "Project description": 0,
                       "Project url": 0, "Priority": 0, "Resolution": 0, "Assignee": 0, "Reporter": 0, "Creator": 0}

# Finds the column index for the Jira sheet and updates the list
for key in Jira_header_index:
    for index, cell in enumerate(jira_sheet[1], 1):
        if key == cell.value:
            Jira_header_index[key] = index
            break  # Break out of the inner loop once a match is found

# Finds the column index for the master sheet and updates the list
for key in master_header_index:
    for index, cell in enumerate(master_sheet[1], 1):
        if key == cell.value:
            master_header_index[key] = index
            break  # Break out of the inner loop once a match is found

# Print header indexes (for debugging)
print("Jira Header Index:", Jira_header_index)
print("Master Header Index:", master_header_index)

# Get the column index for "Issue key" in the Jira sheet
col_num_jira = Jira_header_index["Issue key"]

# Get the column index for "Issue key" in the master sheet
col_num_master = master_header_index["Issue key"]

Jira_letter = get_column_letter(col_num_jira)
master_letter = get_column_letter(col_num_master)

col_jira = jira_sheet[Jira_letter]
col_master = master_sheet[master_letter]

# Find items in Jira that are not in the master sheet
list_a = [cell.value for cell in col_jira if cell.value]
list_b = [cell.value for cell in col_master if cell.value]
not_in_list_b = [(item, i+1) for i, item in enumerate(list_a) if item not in list_b]

# Print items not in the master sheet
print("Items not in Master Sheet:")

# Copy data from Jira sheet to master sheet
for item, row_number in not_in_list_b:
    print(item, row_number)
    data_to_copy = [jira_sheet.cell(row=row_number, column=col).value for col in [ 
        Jira_header_index["Issue key"], 
        Jira_header_index["Updated"],
        Jira_header_index["Custom field (Responsible Team)"],
        Jira_header_index["Summary"], Jira_header_index["Status"], Jira_header_index["Issue id"],Jira_header_index["Issue Type"], 
        Jira_header_index["Custom field (Preventive Action Category)"],Jira_header_index["Project key"],
        Jira_header_index["Project name"],Jira_header_index["Project type"],Jira_header_index["Project lead"], 
        Jira_header_index["Project description"],Jira_header_index["Project url"],Jira_header_index["Priority"],
        Jira_header_index["Resolution"], Jira_header_index["Assignee"],Jira_header_index["Reporter"],
        Jira_header_index["Creator"], 
    ]]

    next_row = output_sheet.max_row + 1

    # Write data to specific columns in the next empty row
    for i, col_name in enumerate(master_header_index.keys()):
        output_sheet.cell(row=next_row, column=master_header_index[col_name], value=data_to_copy[i])
        
        output_sheet.cell(row=next_row, column=1, value=current_date)
    
        next_row = output_sheet.max_row + 1

    # Write data to specific columns in the next empty row
    for i, col_name in enumerate(master_header_index.keys()):
        master_sheet.cell(row=next_row, column=master_header_index[col_name], value=data_to_copy[i])
        
        output_sheet.cell(row=next_row, column=1, value=current_date)
        
# Save and close the workbook
output_workbook.save(output_file_path)
output_workbook.close()

